import os.path

import pandas as pd
import numpy as np
from sklearn.model_selection import train_test_split, StratifiedKFold
from sklearn.metrics import classification_report, precision_recall_fscore_support, accuracy_score, roc_auc_score, precision_score, recall_score, f1_score
import torch
import torch.nn as nn
import torch.optim as optim
import random
from imblearn.over_sampling import SMOTE
from torch.utils.data import DataLoader, TensorDataset
import openpyxl
import argparse
from random import randint
import math
import torch.optim.lr_scheduler as lr_scheduler

# task = 'further'
# run_type = 'select'
# data_path = 'data'
# save_base = 'results'
# random_idx = 1
# prop_idx = 0

parser = argparse.ArgumentParser()

parser.add_argument("--task", default='further', type=str)
parser.add_argument("--run_type", default='select', type=str)
parser.add_argument("--data_path", default='data', type=str)
parser.add_argument("--save_base", default='results', type=str)
parser.add_argument("--random_idx", default=0, type=int)
parser.add_argument("--prop_idx", default=0, type=int)
parser.add_argument("--organ", default='Colon', type=str)
parser.add_argument("--large_model", action='store_true', default=False)

args = parser.parse_args()

# 设置随机数种子
def set_random_seed(seed, deterministic=False):
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)
    if deterministic:
        torch.backends.cudnn.deterministic = True
        torch.backends.cudnn.benchmark = False


set_random_seed(41)
# 读取数据

task = args.task
run_type = args.run_type
data_path = args.data_path
save_base = args.save_base
random_idx = args.random_idx
prop_idx = args.prop_idx
organ = args.organ
large_model = args.large_model
positive_samples_name = os.path.join(data_path, organ, task + '_positive_samples_tpms')
negative_samples_name = os.path.join(data_path, organ, task + '_negative_samples_tpms')
if run_type == 'select':
    positive_samples_name = positive_samples_name + '_select.csv'
    negative_samples_name = negative_samples_name + '_select.csv'
elif run_type == 'all' or run_type == 'mask':
    positive_samples_name = positive_samples_name + '_all.csv'
    negative_samples_name = negative_samples_name + '_all.csv'
elif run_type == 'random':
    positive_samples_name = positive_samples_name + '_random_' + str(random_idx).zfill(4) + '.csv'
    negative_samples_name = negative_samples_name + '_random_' + str(random_idx).zfill(4) + '.csv'
elif run_type == 'prop':
    positive_samples_name = positive_samples_name + '_prop_' + str(prop_idx).zfill(4) + '.csv'
    negative_samples_name = negative_samples_name + '_prop_' + str(prop_idx).zfill(4) + '.csv'


save_path = task + '_' + run_type
if run_type == 'random':
    save_path = save_path + '_' + str(random_idx).zfill(4)
if run_type == 'prop':
    save_path = save_path + '_' + str(prop_idx).zfill(4)
os.makedirs(os.path.join(save_base, organ, save_path), exist_ok=True)

print('positive_samples_name', positive_samples_name)
print('negative_samples_name', negative_samples_name)
positive_samples = pd.read_csv(positive_samples_name)
negative_samples = pd.read_csv(negative_samples_name)
# print('positive_samples', positive_samples)
# exit(0)

if run_type == 'mask':
    select_positive_samples = pd.read_csv(positive_samples_name.replace('_all', '_select'))
# 统计正负样本占比
total_samples = len(positive_samples.columns[1:]) + len(negative_samples.columns[1:])
positive_ratio = len(positive_samples.columns[1:]) / total_samples
negative_ratio = len(negative_samples.columns[1:]) / total_samples
baseline_accuracy = positive_ratio

print(f'Baseline accuracy: {baseline_accuracy:.2f}')

# 转置数据，使样本为行，特征为列，并添加标签
positive_samples_transposed = positive_samples.set_index('ID').T
positive_samples_transposed['label'] = 1

negative_samples_transposed = negative_samples.set_index('ID').T
negative_samples_transposed['label'] = 0

# 合并正负样本
all_samples = pd.concat([positive_samples_transposed, negative_samples_transposed])

# 划分训练集和测试集
X = all_samples.drop(columns=['label'])
y = all_samples['label']

# print('X', X.shape)
# print('y', y.shape)
# exit(0)
# 使用StratifiedKFold进行交叉验证
skf = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)



class EarlyStopping:
    def __init__(self, patience=5, verbose=False, delta=0):
        self.patience = patience
        self.verbose = verbose
        self.delta = delta
        self.counter = 0
        self.best_score = None
        self.early_stop = False
        self.val_loss_min = np.Inf

    def __call__(self, val_loss, model):
        score = val_loss

        if self.best_score is None:
            self.best_score = score
            self.save_checkpoint(val_loss, model)
        elif score < self.best_score + self.delta:
            self.counter += 1
            if self.verbose:
                print(f'EarlyStopping counter: {self.counter} out of {self.patience}')
            if self.counter >= self.patience:
                self.early_stop = True
        else:
            self.best_score = score
            self.save_checkpoint(val_loss, model)
            self.counter = 0

    def save_checkpoint(self, val_loss, model):
        '''Saves model when validation loss decrease.'''
        if self.verbose:
            print(f'Validation loss decreased ({self.val_loss_min:.6f} --> {val_loss:.6f}).  Saving model ...')
        torch.save(model.state_dict(), 'checkpoint.pth')
        self.val_loss_min = val_loss

class TransformerModel(nn.Module):
    def __init__(self, input_dim):
        super(TransformerModel, self).__init__()
        # self.dense1 = nn.Linear(input_dim, 64)
        self.dense1_q = nn.Linear(input_dim, 64)
        self.dense1_k = nn.Linear(input_dim, 64)
        self.dense1_v = nn.Linear(input_dim, 64)
        self.attention = nn.MultiheadAttention(embed_dim=64, num_heads=2)
        self.layer_norm = nn.LayerNorm(64)
        self.dense2 = nn.Linear(64, 32)
        self.dropout = nn.Dropout(0.5)
        self.output = nn.Linear(32, 1)
        self.sigmoid = nn.Sigmoid()

    def forward(self, x):
        # x = torch.relu(self.dense1(x))
        # x = x.unsqueeze(0)  # Adding the sequence dimension
        q = torch.relu(self.dense1_q(x))
        q = q.unsqueeze(0)  # Adding the sequence dimension
        k = torch.relu(self.dense1_k(x))
        k = k.unsqueeze(0)  # Adding the sequence dimension
        v = torch.relu(self.dense1_v(x))
        v = v.unsqueeze(0)  # Adding the sequence dimension
        # x, _ = self.attention(x, x, x)
        x, _ = self.attention(q, k, v)
        x = x.squeeze(0)  # Removing the sequence dimension
        x = self.layer_norm(x)
        x = torch.relu(self.dense2(x))
        x = self.dropout(x)
        x = self.sigmoid(self.output(x))
        return x

class TransformerModelLarge(nn.Module):
    def __init__(self, input_dim):
        super(TransformerModelLarge, self).__init__()
        # self.dense1 = nn.Linear(input_dim, 64)
        dim = 512
        self.head = nn.Linear(input_dim, dim)
        self.dense1_q = nn.Linear(dim, dim)
        self.dense1_k = nn.Linear(dim, dim)
        self.dense1_v = nn.Linear(dim, dim)
        self.attention1 = nn.MultiheadAttention(embed_dim=dim, num_heads=4)
        self.layer_norm1 = nn.LayerNorm(dim)
        self.dense1 = nn.Linear(dim, dim)
        self.dropout1 = nn.Dropout(0.5)
        self.output1 = nn.Linear(dim, dim)

        self.dense2_q = nn.Linear(dim, dim)
        self.dense2_k = nn.Linear(dim, dim)
        self.dense2_v = nn.Linear(dim, dim)
        self.attention2 = nn.MultiheadAttention(embed_dim=dim, num_heads=4)
        self.layer_norm2 = nn.LayerNorm(dim)
        self.dense2 = nn.Linear(dim, dim)
        self.dropout2 = nn.Dropout(0.5)
        self.output2 = nn.Linear(dim, dim)


        self.dense3_q = nn.Linear(dim, dim)
        self.dense3_k = nn.Linear(dim, dim)
        self.dense3_v = nn.Linear(dim, dim)
        self.attention3 = nn.MultiheadAttention(embed_dim=dim, num_heads=4)
        self.layer_norm3 = nn.LayerNorm(dim)
        self.dense3 = nn.Linear(dim, dim)
        self.dropout3 = nn.Dropout(0.5)
        self.output3 = nn.Linear(dim, 1)


        self.sigmoid = nn.Sigmoid()

    def forward(self, x, tau=0.0001):
        x = torch.relu(self.head(x))
        x_identity = x.clone()
        q = torch.relu(self.dense1_q(x))
        q = q.unsqueeze(0)  # Adding the sequence dimension
        k = torch.relu(self.dense1_k(x))
        k = k.unsqueeze(0)  # Adding the sequence dimension
        v = torch.relu(self.dense1_v(x))
        v = v.unsqueeze(0)  # Adding the sequence dimension
        # x, _ = self.attention(x, x, x)
        x, _ = self.attention1(q, k, v)
        x = x.squeeze(0)  # Removing the sequence dimension
        x = self.layer_norm1(x)
        x = torch.relu(self.dense1(x))
        x = self.dropout1(x)
        x = torch.relu(self.output1(x) + x_identity)

        x_identity = x.clone()
        q = torch.relu(self.dense2_q(x))
        q = q.unsqueeze(0)  # Adding the sequence dimension
        k = torch.relu(self.dense2_k(x))
        k = k.unsqueeze(0)  # Adding the sequence dimension
        v = torch.relu(self.dense2_v(x))
        v = v.unsqueeze(0)  # Adding the sequence dimension
        # x, _ = self.attention(x, x, x)
        x, _ = self.attention2(q, k, v)
        x = x.squeeze(0)  # Removing the sequence dimension
        x = self.layer_norm2(x)
        x = torch.relu(self.dense2(x))
        x = self.dropout2(x)
        x = torch.relu(self.output2(x) + x_identity)

        q = torch.relu(self.dense3_q(x))
        q = q.unsqueeze(0)  # Adding the sequence dimension
        k = torch.relu(self.dense3_k(x))
        k = k.unsqueeze(0)  # Adding the sequence dimension
        v = torch.relu(self.dense3_v(x))
        v = v.unsqueeze(0)  # Adding the sequence dimension
        # x, _ = self.attention(x, x, x)
        x, _ = self.attention3(q, k, v)
        x = x.squeeze(0)  # Removing the sequence dimension
        x = self.layer_norm3(x)
        x = torch.relu(self.dense3(x))
        x = self.dropout3(x)
        x = self.sigmoid(self.output3(x))

        return x

def gumbel_softmax(x, dim, tau):
    gumbels = torch.rand_like(x)
    while bool((gumbels == 0).sum() > 0):
        gumbels = torch.rand_like(x)

    gumbels = -(-gumbels.log()).log()
    gumbels = (x + gumbels) / tau
    x = gumbels.softmax(dim)

    return x
class TransformerModelMask(nn.Module):
    def __init__(self, input_dim):
        super(TransformerModelMask, self).__init__()
        # self.dense1 = nn.Linear(input_dim, 64)
        self.mask = nn.Parameter(torch.rand(1, input_dim, 2))
        self.dense1_q = nn.Linear(input_dim, 64)
        self.dense1_k = nn.Linear(input_dim, 64)
        self.dense1_v = nn.Linear(input_dim, 64)
        self.attention = nn.MultiheadAttention(embed_dim=64, num_heads=2)
        self.layer_norm = nn.LayerNorm(64)
        self.dense2 = nn.Linear(64, 32)
        self.dropout = nn.Dropout(0.5)
        self.output = nn.Linear(32, 1)
        self.sigmoid = nn.Sigmoid()

    def forward(self, x, tau=0.0001):
        # x = torch.relu(self.dense1(x))
        # x = x.unsqueeze(0)  # Adding the sequence dimension
        if self.training == True:
            index_num = randint(0, 1)
            if index_num == 0:
                mask = gumbel_softmax(self.mask, 2, tau)[:, :, 0]
            else:
                mask = gumbel_softmax(self.mask, 2, tau)[:, :, 1]
                mask = 1 - mask
        else:
            mask = gumbel_softmax(self.mask, 2, tau)
            mask = (mask[:, :, 0] > mask[:, :, 1]).float()

        x = x * mask
        q = torch.relu(self.dense1_q(x))
        q = q.unsqueeze(0)  # Adding the sequence dimension
        k = torch.relu(self.dense1_k(x))
        k = k.unsqueeze(0)  # Adding the sequence dimension
        v = torch.relu(self.dense1_v(x))
        v = v.unsqueeze(0)  # Adding the sequence dimension
        # x, _ = self.attention(x, x, x)
        x, _ = self.attention(q, k, v)
        x = x.squeeze(0)  # Removing the sequence dimension
        x = self.layer_norm(x)
        x = torch.relu(self.dense2(x))
        x = self.dropout(x)
        x = self.sigmoid(self.output(x))
        return x, mask


class TransformerModelMaskLarge(nn.Module):
    def __init__(self, input_dim):
        super(TransformerModelMaskLarge, self).__init__()
        # self.dense1 = nn.Linear(input_dim, 64)
        dim = 512
        self.mask = nn.Parameter(torch.rand(1, input_dim, 2))
        self.head = nn.Linear(input_dim, dim)
        self.dense1_q = nn.Linear(dim, dim)
        self.dense1_k = nn.Linear(dim, dim)
        self.dense1_v = nn.Linear(dim, dim)
        self.attention1 = nn.MultiheadAttention(embed_dim=dim, num_heads=4)
        self.layer_norm1 = nn.LayerNorm(dim)
        self.dense1 = nn.Linear(dim, dim)
        self.dropout1 = nn.Dropout(0.5)
        self.output1 = nn.Linear(dim, dim)

        self.dense2_q = nn.Linear(dim, dim)
        self.dense2_k = nn.Linear(dim, dim)
        self.dense2_v = nn.Linear(dim, dim)
        self.attention2 = nn.MultiheadAttention(embed_dim=dim, num_heads=4)
        self.layer_norm2 = nn.LayerNorm(dim)
        self.dense2 = nn.Linear(dim, dim)
        self.dropout2 = nn.Dropout(0.5)
        self.output2 = nn.Linear(dim, dim)


        self.dense3_q = nn.Linear(dim, dim)
        self.dense3_k = nn.Linear(dim, dim)
        self.dense3_v = nn.Linear(dim, dim)
        self.attention3 = nn.MultiheadAttention(embed_dim=dim, num_heads=4)
        self.layer_norm3 = nn.LayerNorm(dim)
        self.dense3 = nn.Linear(dim, dim)
        self.dropout3 = nn.Dropout(0.5)
        self.output3 = nn.Linear(dim, 1)


        self.sigmoid = nn.Sigmoid()

    def forward(self, x, tau=0.0001):
        # x = torch.relu(self.dense1(x))
        # x = x.unsqueeze(0)  # Adding the sequence dimension
        if self.training == True:
            index_num = randint(0, 1)
            if index_num == 0:
                mask = gumbel_softmax(self.mask, 2, tau)[:, :, 0]
            else:
                mask = gumbel_softmax(self.mask, 2, tau)[:, :, 1]
                mask = 1 - mask
        else:
            mask = gumbel_softmax(self.mask, 2, tau)
            mask = (mask[:, :, 0] > mask[:, :, 1]).float()

        x = x * mask
        x = torch.relu(self.head(x))
        x_identity = x.clone()
        q = torch.relu(self.dense1_q(x))
        q = q.unsqueeze(0)  # Adding the sequence dimension
        k = torch.relu(self.dense1_k(x))
        k = k.unsqueeze(0)  # Adding the sequence dimension
        v = torch.relu(self.dense1_v(x))
        v = v.unsqueeze(0)  # Adding the sequence dimension
        # x, _ = self.attention(x, x, x)
        x, _ = self.attention1(q, k, v)
        x = x.squeeze(0)  # Removing the sequence dimension
        x = self.layer_norm1(x)
        x = torch.relu(self.dense1(x))
        x = self.dropout1(x)
        x = torch.relu(self.output1(x) + x_identity)

        x_identity = x.clone()
        q = torch.relu(self.dense2_q(x))
        q = q.unsqueeze(0)  # Adding the sequence dimension
        k = torch.relu(self.dense2_k(x))
        k = k.unsqueeze(0)  # Adding the sequence dimension
        v = torch.relu(self.dense2_v(x))
        v = v.unsqueeze(0)  # Adding the sequence dimension
        # x, _ = self.attention(x, x, x)
        x, _ = self.attention2(q, k, v)
        x = x.squeeze(0)  # Removing the sequence dimension
        x = self.layer_norm2(x)
        x = torch.relu(self.dense2(x))
        x = self.dropout2(x)
        x = torch.relu(self.output2(x) + x_identity)

        q = torch.relu(self.dense3_q(x))
        q = q.unsqueeze(0)  # Adding the sequence dimension
        k = torch.relu(self.dense3_k(x))
        k = k.unsqueeze(0)  # Adding the sequence dimension
        v = torch.relu(self.dense3_v(x))
        v = v.unsqueeze(0)  # Adding the sequence dimension
        # x, _ = self.attention(x, x, x)
        x, _ = self.attention3(q, k, v)
        x = x.squeeze(0)  # Removing the sequence dimension
        x = self.layer_norm3(x)
        x = torch.relu(self.dense3(x))
        x = self.dropout3(x)
        x = self.sigmoid(self.output3(x))

        return x, mask

def calculate_noise_levels(data):
    """
    计算每个属性的噪声水平（标准差）。

    参数:
    data (np.ndarray): 输入数据，形状为 (num_samples, num_features)

    返回:
    np.ndarray: 每个属性的标准差
    """
    return np.std(data, axis=0)


def add_noise(data, noise_levels):
    """
    为每个属性添加噪声。噪声水平根据每个属性的尺度来划分。

    参数:
    data (np.ndarray): 输入数据，形状为 (num_samples, num_features)
    noise_levels (list or np.ndarray): 包含每个属性噪声标准差的列表或数组

    返回:
    np.ndarray: 添加了噪声的数据
    """
    # 确保noise_levels是一个np.ndarray
    noise_levels = np.array(noise_levels)

    # 生成与data形状相同的噪声
    noise = np.random.normal(0, noise_levels, data.shape)

    # 返回添加噪声后的数据
    return data + noise


def data_augmentation(X_train, y_train, augmentation_factor=4):
    augmented_X = []
    augmented_y = []
    noise_levels = calculate_noise_levels(X_train)
    for j in range(augmentation_factor):
        for i in range(X_train.shape[0]):
            if j != 0:
                new_sample = add_noise(X_train.iloc[i].values, noise_levels)
            else:
                new_sample = X_train.iloc[i].values
            augmented_X.append(new_sample)
            augmented_y.append(y_train.iloc[i])

    augmented_X = np.array(augmented_X)
    augmented_y = np.array(augmented_y)

    augmented_X = pd.DataFrame(augmented_X, columns=X_train.columns)
    augmented_y = pd.Series(augmented_y)

    return augmented_X, augmented_y


def resample_data(X, y):
    smote = SMOTE(random_state=42)
    X_res, y_res = smote.fit_resample(X, y)
    return X_res, y_res

# 在训练过程中打印每个epoch的准确度
class AccuracyHistory:
    def __init__(self, X_test, y_test):
        self.X_test = X_test
        self.y_test = y_test
        # self.X_test = torch.tensor(self.X_test.values, dtype=torch.float32).cuda()

    def on_epoch_end(self, model, epoch, auc=True):
        model.eval()
        with torch.no_grad():
            # outputs = model(self.X_test)
            outputs = model(torch.tensor(self.X_test.values, dtype=torch.float32).cuda())
            if auc == True:
                outputs = outputs[:,0].cpu().numpy()
                unique_classes = np.unique(self.y_test)
                if len(unique_classes) > 1:
                    auc_results = roc_auc_score(self.y_test, outputs)
                else:
                    auc_results = 0
                # auc_results = roc_auc_score(self.y_test, outputs)
                print(f'Epoch {epoch + 1}, Test auc: {auc_results:.2f}')
                return auc_results
            else:
                predictions = (outputs > 0.5).int().cpu().numpy()
                test_acc = accuracy_score(self.y_test, predictions)
                print(f'Epoch {epoch + 1}, Test accuracy: {test_acc:.2f}')
                return test_acc
    def on_epoch_end_mask(self, model, epoch, auc=True):
        model.eval()
        with torch.no_grad():
            outputs, mask = model(torch.tensor(self.X_test.values, dtype=torch.float32).cuda())
            if auc == True:
                outputs = outputs[:,0].cpu().numpy()
                unique_classes = np.unique(self.y_test)
                if len(unique_classes) > 1:
                    auc_results = roc_auc_score(self.y_test, outputs)
                else:
                    auc_results = 0
                # auc_results = roc_auc_score(self.y_test, outputs)
                print(f'Epoch {epoch + 1}, Test auc: {auc_results:.2f}, Sparse: {torch.mean(mask)}')
                return auc_results
            else:
                predictions = (outputs > 0.5).int().cpu().numpy()
                test_acc = accuracy_score(self.y_test, predictions)
                print(f'Epoch {epoch + 1}, Test accuracy: {test_acc:.2f}, Sparse: {torch.mean(mask)}')
                return test_acc

def print_classification_reports(model, X_val, y_val, X_test, y_test):
    model.eval()
    with torch.no_grad():
        y_val_pred = model(torch.tensor(X_val.values, dtype=torch.float32).cuda())
        y_val_pred = y_val_pred[:,0]
        y_val_pred_classes = (y_val_pred > 0.5).int().cpu().numpy()

        y_val_pred = y_val_pred.cpu().numpy()

        y_test_pred = model(torch.tensor(X_test.values, dtype=torch.float32).cuda())
        y_test_pred = y_test_pred[:, 0]
        y_test_pred_classes = (y_test_pred > 0.5).int().cpu().numpy()

        y_test_pred =  y_test_pred.cpu().numpy()

    acc = accuracy_score(y_val, y_val_pred_classes)
    unique_classes = np.unique(y_val)
    if len(unique_classes) > 1:
        auc = roc_auc_score(y_val, y_val_pred)
    else:
        auc = np.nan
    # auc = roc_auc_score(y_test, y_test_pred)
    precision = precision_score(y_val, y_val_pred_classes, average='binary')
    recall = recall_score(y_val, y_val_pred_classes, average='binary')
    f1 = f1_score(y_val, y_val_pred_classes, average='binary')

    print('valid: auc:' + "{:.4f}".format(auc) + ' f1:' + "{:.4f}".format(f1) + ' acc:' + "{:.4f}".format(acc) + ' prec:' + "{:.4f}".format(precision) + ' recall:' + "{:.4f}".format(recall) )

    acc = accuracy_score(y_test, y_test_pred_classes)
    unique_classes = np.unique(y_test)
    if len(unique_classes) > 1:
        auc = roc_auc_score(y_test, y_test_pred)
    else:
        auc = np.nan
    # auc = roc_auc_score(y_test, y_test_pred)
    precision = precision_score(y_test, y_test_pred_classes, average='binary')
    recall = recall_score(y_test, y_test_pred_classes, average='binary')
    f1 = f1_score(y_test, y_test_pred_classes, average='binary')

    print('test: auc:' + "{:.4f}".format(auc) + ' f1:' + "{:.4f}".format(f1) + ' acc:' + "{:.4f}".format(
        acc) + ' prec:' + "{:.4f}".format(precision) + ' recall:' + "{:.4f}".format(recall))

    return auc, f1, acc, precision, recall

def print_classification_reports_mask(model, X_val, y_val, X_test, y_test):
    model.eval()
    print('X_test', X_test.shape)
    print('X_val', X_val.shape)
    with torch.no_grad():
        y_val_pred, mask_val = model(torch.tensor(X_val.values, dtype=torch.float32).cuda())
        y_val_pred = y_val_pred[:,0]
        y_val_pred_classes = (y_val_pred > 0.5).int().cpu().numpy()

        y_val_pred = y_val_pred.cpu().numpy()

        y_test_pred, mask_test = model(torch.tensor(X_test.values, dtype=torch.float32).cuda())
        y_test_pred = y_test_pred[:, 0]
        y_test_pred_classes = (y_test_pred > 0.5).int().cpu().numpy()

        y_test_pred =  y_test_pred.cpu().numpy()

    acc = accuracy_score(y_val, y_val_pred_classes)
    unique_classes = np.unique(y_val)
    if len(unique_classes) > 1:
        auc = roc_auc_score(y_val, y_val_pred)
    else:
        auc = np.nan
    # auc = roc_auc_score(y_test, y_test_pred)
    precision = precision_score(y_val, y_val_pred_classes, average='binary')
    recall = recall_score(y_val, y_val_pred_classes, average='binary')
    f1 = f1_score(y_val, y_val_pred_classes, average='binary')

    print('valid: auc:' + "{:.4f}".format(auc) + ' f1:' + "{:.4f}".format(f1) + ' acc:' + "{:.4f}".format(acc) + ' prec:' + "{:.4f}".format(precision) + ' recall:' + "{:.4f}".format(recall) + ' mask:' + "{:.4f}".format(torch.mean(mask_val)))

    acc = accuracy_score(y_test, y_test_pred_classes)
    unique_classes = np.unique(y_test)
    if len(unique_classes) > 1:
        auc = roc_auc_score(y_test, y_test_pred)
    else:
        auc = np.nan
    # auc = roc_auc_score(y_test, y_test_pred)
    precision = precision_score(y_test, y_test_pred_classes, average='binary')
    recall = recall_score(y_test, y_test_pred_classes, average='binary')
    f1 = f1_score(y_test, y_test_pred_classes, average='binary')

    print('test: auc:' + "{:.4f}".format(auc) + ' f1:' + "{:.4f}".format(f1) + ' acc:' + "{:.4f}".format(
        acc) + ' prec:' + "{:.4f}".format(precision) + ' recall:' + "{:.4f}".format(recall)+ ' mask:' + "{:.4f}".format(torch.mean(mask_test)))

    return auc, f1, acc, precision, recall

def transfor_format(X_train_aug, X_val, y_train_aug, y_val):

    X_train_aug = X_train_aug.apply(pd.to_numeric, errors='coerce')
    X_train_aug = X_train_aug.fillna(0)  # 填充NaN值为0

    X_val = X_val.apply(pd.to_numeric, errors='coerce')
    X_val = X_val.fillna(0)  # 填充NaN值为0


    y_train_aug = y_train_aug.apply(pd.to_numeric, errors='coerce')
    y_val = y_val.apply(pd.to_numeric, errors='coerce')
    return X_train_aug, X_val, y_train_aug, y_val

if run_type != 'mask':

    # 存储每个测试集的分类报告
    test_reports = []
    acc_test = 0
    prec_test = 0
    rec_test = 0
    f1_test = 0
    auc_test = 0
    auc_count = 0
    # 重新划分训练集、验证集和测试集
    count = 0
    results_l = []
    for train_index, test_index in skf.split(X, y):
        best_model = None
        best_acc = -1

        set_random_seed(42)  # 在每次循环中重新设置随机种子

        X_train, X_test = X.iloc[train_index], X.iloc[test_index]
        y_train, y_test = y.iloc[train_index], y.iloc[test_index]

        X_test = X_test.apply(pd.to_numeric, errors='coerce')
        X_test = X_test.fillna(0)  # 填充NaN值为0
        y_test = y_test.apply(pd.to_numeric, errors='coerce')

        # 对训练集进行重采样
        # X_train_res, y_train_res = resample_data(X_train, y_train)
        X_train_res = X_train
        y_train_res = y_train
        # # 对训练集进行数据增强
        # X_train_aug, y_train_aug = data_augmentation(X_train_res, y_train_res)
        #
        # # 从训练集中划分验证集
        # X_train_aug, X_val, y_train_aug, y_val = train_test_split(X_train_aug, y_train_aug, test_size=0.2,
        #                                                           stratify=y_train_aug, random_state=42)

        # 从训练集中划分验证集
        X_train_aug, X_val, y_train_aug, y_val = train_test_split(X_train_res, y_train_res, test_size=0.2,
                                                                  stratify=y_train_res, random_state=42)

        # # 对训练集进行数据增强
        # X_train_aug, y_train_aug = data_augmentation(X_train_aug, y_train_aug)



        # 构建带有Transformer模块的神经网络模型
        if large_model == False:
            model = TransformerModel(X_train_aug.shape[1])
        elif large_model == True:
            model = TransformerModelLarge(X_train_aug.shape[1])

        model = model.cuda()

        criterion = nn.BCELoss()
        optimizer = optim.Adam(model.parameters(), lr=0.001)

        if organ == 'Multitask':
            X_train_aug, X_val, y_train_aug, y_val = transfor_format(X_train_aug, X_val, y_train_aug, y_val)

        # 训练模型
        accuracy_history = AccuracyHistory(X_val, y_val)
        train_loader = DataLoader(TensorDataset(torch.tensor(X_train_aug.values, dtype=torch.float32),
                                                torch.tensor(y_train_aug.values, dtype=torch.float32)), batch_size=16,
                                  shuffle=True, pin_memory=True)

        # early_stopping = EarlyStopping(patience=5, verbose=True)
        epoch_all =50
        if organ == 'Multitask':
            epoch_all = 300
        if task == 'tumor':
            epoch_all=5
        lf = lambda x: ((1 + math.cos(x * math.pi / epoch_all)) / 2) * (1 - 0.0001) + 0.0001
        scheduler = lr_scheduler.LambdaLR(optimizer, lr_lambda=lf)
        for epoch in range(epoch_all):
            model.train()
            accu_num = 0
            sample_num = 0
            for batch_X, batch_y in train_loader:
                optimizer.zero_grad()
                batch_X = batch_X.cuda()
                batch_y = batch_y.cuda()
                outputs = model(batch_X)
                loss = criterion(outputs[:,0], batch_y)
                loss.backward()
                optimizer.step()
                pred_classes = (outputs > 0.5).long()
                accu_num += (pred_classes[:,0]==batch_y.long()).sum()
                sample_num += batch_X.shape[0]
            if organ == 'Multitask':
                scheduler.step()
                current_lr = scheduler.get_last_lr()[0]
                print(f"Epoch {epoch + 1}/{epoch_all}, Learning Rate: {current_lr}")
            print(f"Epoch {epoch + 1}, Train accuracy: {accu_num.item() / sample_num:.2f}")

            val_acc = accuracy_history.on_epoch_end(model, epoch, auc=True)
            if val_acc > best_acc:
                best_acc = val_acc
                # best_model = model
                torch.save(model.state_dict(), os.path.join(save_base, organ, save_path, 'best_model_'+ str(count).zfill(4) + '.pth'))

        if large_model == False:
            best_model = TransformerModel(X_train_aug.shape[1])
        else:
            print('use TransformerModelLarge')
            best_model = TransformerModelLarge(X_train_aug.shape[1])
        best_model.load_state_dict(torch.load(os.path.join(save_base, organ, save_path, 'best_model_'+ str(count).zfill(4) + '.pth')))
        best_model = best_model.cuda()

        auc, f1, acc, precision, recall = print_classification_reports(best_model, X_val, y_val, X_test, y_test)
        acc_test += acc
        prec_test += precision
        rec_test += recall
        f1_test += f1
        if np.isnan(auc) == False:
            auc_test += auc
            auc_count += 1

        results = ['Fold_' + str(count).zfill(2), auc, f1, acc, precision, recall]
        results_l.append(results)

        count += 1

    print("Test Mean AUC:", auc_test / auc_count)
    print("Test Mean F1-Score:", f1_test / 5)
    print("Test Mean Accuracy:", acc_test / 5)
    print("Test Mean Precision:", prec_test / 5)
    print("Test Mean Recall:", rec_test / 5)

    results_l.append(['Average', auc_test / auc_count, f1_test / 5, acc_test / 5, prec_test / 5, rec_test / 5])



    cells = ['', 'AUC', 'F1', 'ACC', 'PREC', 'REC']
    results_l.insert(0, cells)
    wb = openpyxl.Workbook()
    ws = wb.active

    for row_index, data_list in enumerate(results_l, start=1):
        for col_index, value in enumerate(data_list, start=1):
            if isinstance(value, (float, int)):  # 仅格式化数值类型
                value = "{:.4f}".format(value)
            ws.cell(row=row_index, column=col_index, value=value)
    wb.save(os.path.join(save_base, organ, save_path, 'results.xlsx'))
else:

    # 存储每个测试集的分类报告
    test_reports = []
    acc_test = 0
    prec_test = 0
    rec_test = 0
    f1_test = 0
    auc_test = 0
    auc_count = 0
    # 重新划分训练集、验证集和测试集
    count = 0
    results_l = []
    for train_index, test_index in skf.split(X, y):
        best_model = None
        best_acc = -1

        set_random_seed(42)  # 在每次循环中重新设置随机种子

        X_train, X_test = X.iloc[train_index], X.iloc[test_index]
        y_train, y_test = y.iloc[train_index], y.iloc[test_index]

        X_test = X_test.apply(pd.to_numeric, errors='coerce')
        X_test = X_test.fillna(0)  # 填充NaN值为0
        y_test = y_test.apply(pd.to_numeric, errors='coerce')

        # 对训练集进行重采样
        # X_train_res, y_train_res = resample_data(X_train, y_train)
        X_train_res = X_train
        y_train_res = y_train
        # # 对训练集进行数据增强
        # X_train_aug, y_train_aug = data_augmentation(X_train_res, y_train_res)
        #
        # # 从训练集中划分验证集
        # X_train_aug, X_val, y_train_aug, y_val = train_test_split(X_train_aug, y_train_aug, test_size=0.2,
        #                                                           stratify=y_train_aug, random_state=42)

        # 从训练集中划分验证集
        X_train_aug, X_val, y_train_aug, y_val = train_test_split(X_train_res, y_train_res, test_size=0.2,
                                                                  stratify=y_train_res, random_state=42)

        # # 对训练集进行数据增强
        # X_train_aug, y_train_aug = data_augmentation(X_train_aug, y_train_aug)

        # 构建带有Transformer模块的神经网络模型
        # model = TransformerModelMask(X_train_aug.shape[1])
        if large_model == False:
            model = TransformerModelMask(X_train_aug.shape[1])
        elif large_model == True:
            model = TransformerModelMaskLarge(X_train_aug.shape[1])
        model = model.cuda()

        criterion = nn.BCELoss()
        optimizer = optim.Adam(model.parameters(), lr=0.001)

        if organ == 'Multitask':
            X_train_aug, X_val, y_train_aug, y_val = transfor_format(X_train_aug, X_val, y_train_aug, y_val)

        # 训练模型
        accuracy_history = AccuracyHistory(X_val, y_val)
        train_loader = DataLoader(TensorDataset(torch.tensor(X_train_aug.values, dtype=torch.float32),
                                                torch.tensor(y_train_aug.values, dtype=torch.float32)), batch_size=16,
                                  shuffle=True, pin_memory=True)

        # early_stopping = EarlyStopping(patience=5, verbose=True)
        if len(train_loader) < 16:
            divide_base = 1
        else:
            divide_base = len(train_loader) // 16
        epoch_all = 20 * 300 // divide_base
        print('epoch_all', epoch_all)
        print_val = epoch_all // 10
        if epoch_all < 300:
            epoch_all = 300
        # print('select_positive_samples', select_positive_samples.shape)
        # print('positive_samples', positive_samples.shape)
        tgt_sparse = select_positive_samples.shape[0] / positive_samples.shape[0]
        print('tgt_sparse', tgt_sparse)
        # if task == 'tumor':
        #     epoch_all = 5

        def exponential_decay(epoch, total_epochs, tau_initial=0.4, tau_final=0.0001):
            tau = tau_initial * (tau_final / tau_initial) ** (epoch / total_epochs)
            return tau

        lf = lambda x: ((1 + math.cos(x * math.pi / epoch_all)) / 2) * (1 - 0.0001) + 0.0001
        scheduler = lr_scheduler.LambdaLR(optimizer, lr_lambda=lf)

        for epoch in range(epoch_all):
            model.train()
            accu_num = 0
            sample_num = 0
            for batch_X, batch_y in train_loader:
                optimizer.zero_grad()
                batch_X = batch_X.cuda()
                batch_y = batch_y.cuda()
                tau = exponential_decay(epoch, epoch_all, 1.0, 0.0001)
                outputs, mask = model(batch_X, tau)
                loss = criterion(outputs[:,0], batch_y) + torch.abs(torch.mean(mask) - tgt_sparse) * 10.0
                loss.backward()
                optimizer.step()
                pred_classes = (outputs > 0.5).long()
                accu_num += (pred_classes[:, 0] == batch_y.long()).sum()
                sample_num += batch_X.shape[0]
            if organ == 'Multitask':
                scheduler.step()
                current_lr = scheduler.get_last_lr()[0]
                print(f"Epoch {epoch + 1}/{epoch_all}, Learning Rate: {current_lr}")
            print(f"Epoch {epoch + 1}, Train accuracy: {accu_num.item() / sample_num:.2f}")
            if epoch % print_val == 0:
                val_acc = accuracy_history.on_epoch_end_mask(model, epoch, auc=True)
            # if val_acc > best_acc:
            #     best_acc = val_acc
            #     # best_model = model
        # print(os.path.join(save_base, organ, save_path, 'best_model_' + str(count).zfill(4) + '.pth'))
        torch.save(model.state_dict(),
                   os.path.join(save_base, organ, save_path, 'best_model_' + str(count).zfill(4) + '.pth'))

        if large_model == False:
            best_model = TransformerModelMask(X_train_aug.shape[1])
        else:
            best_model = TransformerModelMaskLarge(X_train_aug.shape[1])
            print('use TransformerModelMaskLarge')
        best_model.load_state_dict(
            torch.load(os.path.join(save_base, organ, save_path, 'best_model_' + str(count).zfill(4) + '.pth')))
        best_model = best_model.cuda()

        auc, f1, acc, precision, recall = print_classification_reports_mask(best_model, X_val, y_val, X_test, y_test)
        acc_test += acc
        prec_test += precision
        rec_test += recall
        f1_test += f1
        if np.isnan(auc) == False:
            auc_test += auc
            auc_count += 1

        results = ['Fold_' + str(count).zfill(2), auc, f1, acc, precision, recall]
        results_l.append(results)

        count += 1

    print("Test Mean AUC:", auc_test / auc_count)
    print("Test Mean F1-Score:", f1_test / 5)
    print("Test Mean Accuracy:", acc_test / 5)
    print("Test Mean Precision:", prec_test / 5)
    print("Test Mean Recall:", rec_test / 5)

    results_l.append(['Average', auc_test / auc_count, f1_test / 5, acc_test / 5, prec_test / 5, rec_test / 5])

    cells = ['', 'AUC', 'F1', 'ACC', 'PREC', 'REC']
    results_l.insert(0, cells)
    wb = openpyxl.Workbook()
    ws = wb.active

    for row_index, data_list in enumerate(results_l, start=1):
        for col_index, value in enumerate(data_list, start=1):
            if isinstance(value, (float, int)):  # 仅格式化数值类型
                value = "{:.4f}".format(value)
            ws.cell(row=row_index, column=col_index, value=value)
    wb.save(os.path.join(save_base, organ, save_path, 'results.xlsx'))