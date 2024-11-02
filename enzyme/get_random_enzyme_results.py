import os.path

import pandas as pd
import openpyxl

organ_names = [
    'Multitask'
]


task_l = ['tumor']
base_result = 'results'

select_xlsx_l = ['tumor_mask_0.0042', 'tumor_mask_0.005', 'tumor_mask_0.02', 'tumor_mask_0.03', 'tumor_mask_0.05']

for organ in organ_names:
    for task in task_l:
        print(task)

        subpath_base = 'random_'
        subtasks=['select_', 'select_in_']
        save_name_l = ['All_random', 'Select_random']
        num_l = [2, 5, 9, 16, 26]
        for num_idx, num in enumerate(num_l):
            for subtask in subtasks:
                xlsx_l = []
                cells = ['', 'AUC', 'F1', 'ACC', 'PREC', 'REC']
                xlsx_l.append(cells)
                metric_sum_l = [[] for i in range(5)]
                for i in range(100):
                    xlsx_file = os.path.join(base_result, organ, task + '_' + subpath_base + subtask + str(num).zfill(4) + '_' + str(i).zfill(4), 'results.xlsx')
                    print(xlsx_file)
                    if os.path.exists(xlsx_file) == True:
                        df = pd.read_excel(xlsx_file)
                        last_row = df.iloc[-1]
                        last_row = last_row[1:]
                        last_row_floats = last_row.astype(float).values
                        for f_idx, f in enumerate(last_row_floats):
                            metric_sum_l[f_idx].append(f)

                        last_row_str = df.iloc[-1]
                        last_row_str[0] = str(i).zfill(4)
                        xlsx_l.append(last_row_str)

                cells = ['Average']
                for metric_sum in metric_sum_l:
                    cells.append(sum(metric_sum)/ len(metric_sum))
                xlsx_l.append(cells)

                wb = openpyxl.Workbook()
                ws = wb.active
                for row_index, data_list in enumerate(xlsx_l, start=1):
                    for col_index, value in enumerate(data_list, start=1):
                        # if isinstance(value, (float, int)):  # 仅格式化数值类型
                        #     value = "{:.4f}".format(value)
                        ws.cell(row=row_index, column=col_index, value=value)
                wb.save(os.path.join(base_result, organ, task + '_' + subpath_base + subtask + str(num).zfill(4) + '.xlsx'))
                print(os.path.join(base_result, organ, task + '_' + subpath_base + subtask + str(num).zfill(4) + '.xlsx'))


        xlsx_l  = []
        xlsx_single_l = []
        cells = ['', 'AUC', 'F1', 'ACC', 'PREC', 'REC']
        xlsx_l.append(cells)
        xlsx_single_l.append(cells)
        for num_idx, num in enumerate(num_l):
            for subtask_idx, subtask in enumerate(subtasks):
                file_path = os.path.join(base_result, organ, task + '_' + subpath_base + subtask + str(num).zfill(4) + '.xlsx')
                df = pd.read_excel(file_path)
                last_row = df.iloc[-1]
                last_row[0] = save_name_l[subtask_idx] + '_' + str(num).zfill(4)
                xlsx_l.append(last_row)
                xlsx_single_l.append(last_row)

            file_path = os.path.join(base_result, organ, select_xlsx_l[num_idx], 'results_merge.xlsx')
            df = pd.read_excel(file_path)
            last_row = df.iloc[-1]
            last_row[0] = 'Select' + '_' + str(num).zfill(4)
            xlsx_l.append(last_row)
            xlsx_single_l.append(last_row)

            wb = openpyxl.Workbook()
            ws = wb.active
            for row_index, data_list in enumerate(xlsx_single_l, start=1):
                for col_index, value in enumerate(data_list, start=1):
                    if isinstance(value, (float, int)):  # 仅格式化数值类型
                        value = "{:.4f}".format(value)
                    ws.cell(row=row_index, column=col_index, value=value)
            wb.save(os.path.join(base_result, organ, task + '_select_ablation_' + str(num).zfill(4) + '.xlsx'))
            print(os.path.join(base_result, organ, task + '_select_ablation_' + str(num).zfill(4) + '.xlsx'))
            xlsx_single_l = []

        wb = openpyxl.Workbook()
        ws = wb.active
        for row_index, data_list in enumerate(xlsx_l, start=1):
            for col_index, value in enumerate(data_list, start=1):
                if isinstance(value, (float, int)):  # 仅格式化数值类型
                    value = "{:.4f}".format(value)
                ws.cell(row=row_index, column=col_index, value=value)
        wb.save(os.path.join(base_result, organ, task + '_select_ablation_' + '.xlsx'))
        print(os.path.join(base_result, organ, task + '_select_ablation_' + '.xlsx'))