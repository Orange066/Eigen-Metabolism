import os.path

import pandas as pd
import openpyxl

organ_names = [
    'Multitask'
]


task_l = ['tumor']
base_result = 'results'

select_xlsx_l = ['tumor_mask_0.0042', 'tumor_mask_0.005', 'tumor_mask_0.02', 'tumor_mask_0.03', 'tumor_mask_0.05']


######## all ########
xlsx_l = []
cells = ['', 'AUC', 'F1', 'ACC', 'PREC', 'REC']
xlsx_l.append(cells)

file_path = './results/Multitask/tumor_mask_0.0042/results_merge.xlsx'
df = pd.read_excel(file_path)
last_row = df.iloc[-1]
last_row[0] = '2'
xlsx_l.append(last_row)

file_path = './results/Multitask/tumor_incre_all_1/results.xlsx'
df = pd.read_excel(file_path)
last_row = df.iloc[-1]
last_row[0] = '5-2'
xlsx_l.append(last_row)

file_path = './results/Multitask/tumor_incre_all_2/results.xlsx'
df = pd.read_excel(file_path)
last_row = df.iloc[-1]
last_row[0] = '9-5'
xlsx_l.append(last_row)

file_path = './results/Multitask/tumor_incre_all_3/results.xlsx'
df = pd.read_excel(file_path)
last_row = df.iloc[-1]
last_row[0] = '16-9'
xlsx_l.append(last_row)

file_path = './results/Multitask/tumor_incre_all_4/results.xlsx'
df = pd.read_excel(file_path)
last_row = df.iloc[-1]
last_row[0] = '26-16'
xlsx_l.append(last_row)

wb = openpyxl.Workbook()
ws = wb.active
for row_index, data_list in enumerate(xlsx_l, start=1):
    for col_index, value in enumerate(data_list, start=1):
        if isinstance(value, (float, int)):  # 仅格式化数值类型
            value = "{:.4f}".format(value)
        ws.cell(row=row_index, column=col_index, value=value)
wb.save('./results/Multitask/tumor_incre_all.xlsx')


######## signle - 2 ########

xlsx_l = []
cells = ['', 'AUC', 'F1', 'ACC', 'PREC', 'REC']
xlsx_l.append(cells)

file_path = './results/Multitask/tumor_mask_0.0042/results_merge.xlsx'
df = pd.read_excel(file_path)
last_row = df.iloc[-1]
last_row[0] = '2'
xlsx_l.append(last_row)

file_path = './results/Multitask/tumor_incre_single_0000_0001/results.xlsx'
df = pd.read_excel(file_path)
last_row = df.iloc[-1]
last_row[0] = '5-2'
xlsx_l.append(last_row)

file_path = './results/Multitask/tumor_incre_single_0000_0002/results.xlsx'
df = pd.read_excel(file_path)
last_row = df.iloc[-1]
last_row[0] = '9-2'
xlsx_l.append(last_row)

file_path = './results/Multitask/tumor_incre_single_0000_0003/results.xlsx'
df = pd.read_excel(file_path)
last_row = df.iloc[-1]
last_row[0] = '16-2'
xlsx_l.append(last_row)

file_path = './results/Multitask/tumor_incre_single_0000_0004/results.xlsx'
df = pd.read_excel(file_path)
last_row = df.iloc[-1]
last_row[0] = '26-2'
xlsx_l.append(last_row)

wb = openpyxl.Workbook()
ws = wb.active
for row_index, data_list in enumerate(xlsx_l, start=1):
    for col_index, value in enumerate(data_list, start=1):
        if isinstance(value, (float, int)):  # 仅格式化数值类型
            value = "{:.4f}".format(value)
        ws.cell(row=row_index, column=col_index, value=value)
wb.save('./results/Multitask/tumor_incre_single_2.xlsx')


######## signle - 4 ########

xlsx_l = []
cells = ['', 'AUC', 'F1', 'ACC', 'PREC', 'REC']
xlsx_l.append(cells)

file_path = './results/Multitask/tumor_mask_0.005/results_merge.xlsx'
df = pd.read_excel(file_path)
last_row = df.iloc[-1]
last_row[0] = '5'
xlsx_l.append(last_row)

file_path = './results/Multitask/tumor_incre_single_0001_0002/results.xlsx'
df = pd.read_excel(file_path)
last_row = df.iloc[-1]
last_row[0] = '9-5'
xlsx_l.append(last_row)

file_path = './results/Multitask/tumor_incre_single_0001_0003/results.xlsx'
df = pd.read_excel(file_path)
last_row = df.iloc[-1]
last_row[0] = '16-5'
xlsx_l.append(last_row)

file_path = './results/Multitask/tumor_incre_single_0001_0004/results.xlsx'
df = pd.read_excel(file_path)
last_row = df.iloc[-1]
last_row[0] = '16-5'
xlsx_l.append(last_row)


wb = openpyxl.Workbook()
ws = wb.active
for row_index, data_list in enumerate(xlsx_l, start=1):
    for col_index, value in enumerate(data_list, start=1):
        if isinstance(value, (float, int)):  # 仅格式化数值类型
            value = "{:.4f}".format(value)
        ws.cell(row=row_index, column=col_index, value=value)
wb.save('./results/Multitask/tumor_incre_single_5.xlsx')


######## signle - 10 ########

xlsx_l = []
cells = ['', 'AUC', 'F1', 'ACC', 'PREC', 'REC']
xlsx_l.append(cells)

file_path = './results/Multitask/tumor_mask_0.02/results_merge.xlsx'
df = pd.read_excel(file_path)
last_row = df.iloc[-1]
last_row[0] = '9'
xlsx_l.append(last_row)

file_path = './results/Multitask/tumor_incre_single_0001_0002/results.xlsx'
df = pd.read_excel(file_path)
last_row = df.iloc[-1]
last_row[0] = '16-9'
xlsx_l.append(last_row)

file_path = './results/Multitask/tumor_incre_single_0001_0003/results.xlsx'
df = pd.read_excel(file_path)
last_row = df.iloc[-1]
last_row[0] = '26-9'
xlsx_l.append(last_row)


wb = openpyxl.Workbook()
ws = wb.active
for row_index, data_list in enumerate(xlsx_l, start=1):
    for col_index, value in enumerate(data_list, start=1):
        if isinstance(value, (float, int)):  # 仅格式化数值类型
            value = "{:.4f}".format(value)
        ws.cell(row=row_index, column=col_index, value=value)
wb.save('./results/Multitask/tumor_incre_single_9.xlsx')


######## signle - 19 ########

xlsx_l = []
cells = ['', 'AUC', 'F1', 'ACC', 'PREC', 'REC']
xlsx_l.append(cells)

file_path = './results/Multitask/tumor_mask_0.03/results_merge.xlsx'
df = pd.read_excel(file_path)
last_row = df.iloc[-1]
last_row[0] = '16'
xlsx_l.append(last_row)

file_path = './results/Multitask/tumor_incre_single_0001_0002/results.xlsx'
df = pd.read_excel(file_path)
last_row = df.iloc[-1]
last_row[0] = '26-16'
xlsx_l.append(last_row)



wb = openpyxl.Workbook()
ws = wb.active
for row_index, data_list in enumerate(xlsx_l, start=1):
    for col_index, value in enumerate(data_list, start=1):
        if isinstance(value, (float, int)):  # 仅格式化数值类型
            value = "{:.4f}".format(value)
        ws.cell(row=row_index, column=col_index, value=value)
wb.save('./results/Multitask/tumor_incre_single_16.xlsx')


######## signle - 26 ########

xlsx_l = []
cells = ['', 'AUC', 'F1', 'ACC', 'PREC', 'REC']
xlsx_l.append(cells)

file_path = './results/Multitask/tumor_mask_0.05/results_merge.xlsx'
df = pd.read_excel(file_path)
last_row = df.iloc[-1]
last_row[0] = '26'
xlsx_l.append(last_row)


wb = openpyxl.Workbook()
ws = wb.active
for row_index, data_list in enumerate(xlsx_l, start=1):
    for col_index, value in enumerate(data_list, start=1):
        if isinstance(value, (float, int)):  # 仅格式化数值类型
            value = "{:.4f}".format(value)
        ws.cell(row=row_index, column=col_index, value=value)
wb.save('./results/Multitask/tumor_incre_single_26.xlsx')